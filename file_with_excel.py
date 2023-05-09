from openpyxl import Workbook

wb = Workbook()

# grab the ative worsheets

ws = wb.active

ws['A1'] = 50

ws.append([1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11])


import datetime


ws['A2'] = datetime.datetime.now().time()

# save the file

wb.save('sample.xlsx')