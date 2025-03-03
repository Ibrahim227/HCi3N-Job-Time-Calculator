import datetime
import os
import tkinter as tk
import webbrowser
from time import strftime
from tkinter import ttk, messagebox, END, BOTH

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from startup_image import mainwindow


# redirect to HCi3N website
def on_click():
    """
    Redirect to HC3N website
    :return: Open a new tab within your default web-browser
    """
    url = "http://www.initiative3n.ne/"
    webbrowser.open_new_tab(url)


####### function for Startup Image
def start_up_image():
    """
    :return: Display the startup image
    """
    win = mainwindow
    win()


### Class Definition
class JobTimeCalculator:
    # Initialize the class
    def __init__(self):
        super().__init__()  # Allows inheriting from the tkinter class object
        # Main window
        self.window = tk.Tk()
        self.window.title("HC3N")
        self.window.resizable(True, True)
        # self.window.bell(displayof=0)
        self.window.geometry()
        self.window.wait_visibility(window=self.window)
        self.window.iconbitmap('images\\logoHCi3N.ico')
        self.window.config(background="lightgrey", highlightthickness=False, relief="groove", border=5)
        self.window.iconposition(x=5, y=1)

        #
        ###

        ######## Main Frame
        self.frame = ttk.Frame(self.window, relief='flat')
        self.frame.pack(side='top', fill=BOTH)

        ################################# Configure First LabelFrame ############################
        # Saving user Information
        self.user_info_frame = ttk.Labelframe(self.frame, text='INFORMATION EMPLOYEE', underline=0)
        self.user_info_frame.grid(row=0, column=0, sticky='news', padx=20, pady=10)

        ################# lateral LabelFrame #################
        self.lateral_label_frame = ttk.LabelFrame(self.frame,
                                                  text="NOM & PRENOM - OBSERVATION - DATE",
                                                  underline=0)
        self.lateral_label_frame.grid(row=0, column=1, sticky='news', padx=20, pady=10)

        # Create name and last name label
        self.first_last_name_label = ttk.Label(self.lateral_label_frame, text='NOM & PRENOM:', background="lightgrey",
                                               underline=0)
        self.first_last_name_label.grid(row=0, column=0)

        # create function_set and last name entry widgets
        self.name_list = sorted(
            ["ALI BETY", "ABDOULAYE MAIZAMA", "VINCENT PARAISO MOUSSA", "Mme RABO MARIA MOHAMED YAROH",
             "BOUKARY ABDOU RAZAK", "ABDOU KASSO",
             "ABOUBACAR DJIMRAOU", "ABOUBA SAIDOU", "IDRISSA CHIPKAOU", "KORAO ABOUBACAR",
             "Mme DOUDOU HALIDOU MAIMOUNA", "GAMATIE BOUBACAR", "Mme ABDOUL NASSER MARIA",
             "RABIATOU HABIBOU", "ALASSANE ABDOU ALMOUSTAPHA", "ALI OUMAROU", "OUSMANE FODI",
             "Mme MOUSTAPHA FOURERA", "Mme MARIAMA AROUNA ANOUAR", "Mme IDRISSA NANA AICHATOU",
             "Mme ABDOURAHAMANE FOURERATOU DIALLO", "Mme DJIBO ZEYNABOU COULIBALY",
             "MAHAMADOU MAHAMANE NAFISSATOU", "Mme AMINATOU MAHAMAN ALTINÉ", "Mme FATIMA ISSA BOUKARI",
             "ABDOU MOUSSA OUSMANE",
             "GOUMAR ALHASSANE", "ABDOULAYE TANKARI AMADOU", "HAMANI TAHIROU SAIDOU RACHID",
             "TRAPSIDA ABDOULAY ALAIN", "BOUBACAR HAMADOU", "IBRAHIM MOUSSA", "IBRAHIM DJIBRILLA",
             "MA AROUF TIDJANI", "LAWAN DARMANE", "ALI OUMAROU", "Mme SALAMATOU AMADOU",
             "BOGARI ZOURKALEINI", "Maman Sani Ibrahim",
             "BABA BAFRAGI BOUBACAR", "HAMA AMADOU", "GN ABOUBACAR OUMAROU KAILOU",
             "GN HAROUNA MAAZOU LEYO", "GN LAOUALI MAAZOU MAMANE", "MOCTAR BACHIR",
             "MALAM ROUFAI MAMAN SANI", "ELHADJI SEYBOU DJIBO", "MOUSTAPHA AHMET",
             "AMADOU BACHIR", "OUSMANE YERIMA YAHAYA", "MAMOUDOU MAHAMAN BACHAR",
             "SALAMATOU SOUMANA MOUSSA", "SAHABI ABDOU", "GN KABIROU ABDOUL MOUMOUNI",
             "GN ADAHIR IDI DJIBAGÉ", "Dr MAHAMADOU ABOUBACAR",
             "AMINA IDRISSA BAGNOU", "ABDOU ADAMOU LILWANI", "ABDOUL WAHABOU ZAKARI DAGOU",
             "ISSA HAMANI ABDOULAYE", "ALI OUSSEINI MOUSTAPHA", "ALI SOUMAILA FOUREIRATOU",
             "YAHAYA RHISSA ZAKARI", "ABDALLAH MAHAMAT YAHAYA", "ABASS ADAM MELLY HADIZA",
             "SOULEY BOUKAR", "HAMIDOU AMANI SOULEYMANE", "KOUNKOUROU AHAMADOU",
             "Mme SEYDOU ABDOULAYE FOUREYRATOU", "GN RABIOU ABDOULAYE WACHEL", "GN ALMOUSTAPHA DJIBAGÉ"])

        self.first_last_name_entry = ttk.Combobox(self.lateral_label_frame, values=self.name_list)
        self.first_last_name_entry.grid(row=0, column=1, ipadx=60)

        # create the observation and date Entry
        self.observation_list_combobox_label = ttk.Label(self.lateral_label_frame, text="OBSERVATION:",
                                                         background="lightgrey",
                                                         underline=0)
        self.observation_list_combobox_label.grid(row=2, column=0)
        self.observation_list_combobox = ttk.Combobox(self.lateral_label_frame,
                                                      values=(sorted(["Sorti(e) pour SIEGE", "Sorti(e) pour ANNEXE-1",
                                                                      "Sorti(e) pour ANNEXE-2", "Réunion", "Atelier",
                                                                      "Mission", "Abscence",
                                                                      "Consultation", "Décès", "Maladie", "Mariage",
                                                                      "Permission", "Congé", "Autres",
                                                                      "Non Préciser", "Congé maternité"])))
        self.observation_list_combobox.grid(row=2, column=1, ipadx=20)

        self.date_entry_label = ttk.Label(self.lateral_label_frame, text="DATE JOUR/MOIS/ANNEE:",
                                          background="lightgrey",
                                          underline=0)
        self.date_entry_label.grid(row=3, column=0)

        # self.date_entry = ttk.Entry(self.lateral_label_frame)
        #
        # self.date_entry.grid(row=3, column=1, ipadx=20)

        ##### Calendar function
        ##### Calendar function
        # def get_selected_date():
        #     self.spinbox.get_date()
        # self.spinbox = DateEntry(self.lateral_label_frame,background='darkblue',foreground='white',borderwidth=2)
        # self.spinbox.grid(row=3, column=1)
        # self.spinbox.bind("<<DateEntrySelected>>", lambda event: get_selected_date)

        self.date_entry = ttk.Entry(self.lateral_label_frame)
        self.date_entry.grid(row=3, column=1)

        for widget in self.lateral_label_frame.winfo_children():
            widget.grid_configure(padx=20, pady=10, sticky="news")
        #
        #
        # Create the title combobox
        self.title = ttk.Label(self.user_info_frame, text='FONCTION:', background="lightgrey", underline=0)
        self.title_combox = ttk.Combobox(self.user_info_frame,
                                         values=(sorted(
                                             ["HAUT-COMMISSAIRE", "SG", "SGA",
                                              "DIRECTRICE DAFC", "SECRETAIRE DE DIRECTION",
                                              "SECRETAIRE DE DIRECTION/BO", "CONSEILLER TECHNIQUE",
                                              "CHEF DE CABINET", "PROTOCOLE", "DIRECTEUR DPSFCI",
                                              "CHEF DIV INFORMATIQUE", "CHEF DIV AFFAIRES FINANCIERE",
                                              "CHEF DIV. MARCHES PUBLIC/DSP",
                                              "GESTIONNAIRE DES CONVENTIONS", "CHEF DIV PATRIMOINE LOGISTIQUE",
                                              "APPELEE SERV CIVIQUE", "STAGIAIRE",
                                              "CHEF DIV RH", "DIRECTEUR DMRC", "CHEF DIV RENFORCEMENT CAPACITE",
                                              "CHEF DIV MOBILISATION SOCIAL", "SECURITÉ RAPPROCHÉ",
                                              "DIRECTEUR DSEC", "CHEF DIV CAPITALISATION", "CHEF DIV SISAN",
                                              "CHEF DIV SUIVI-EVALUATION STATISTIQUE",
                                              "DIRECTEUR DPEP", "CHEF DIV PROGRAMMATION", "COORDINATEUR",
                                              "CHEF SERV CARTOGRAPHY", "CHEF SERV COM/CELLULE NUTRITION",
                                              "CHEF SERV BIO-STATISTIQUE", "ING STATISTICIEN ECONOMISTE P",
                                              "MEDECIN NUTRITIONISTE", "COORDINATRICE Proj NEXUS",
                                              "Resp ADMINISTRATIF FINANCE", "Resp SUIVI-EVALUATION",
                                              "SECRETAIRE COMPTABLE", "CR/TAHOUA", "AT CRi3N NIAMEY",
                                              "ASSISTANTE SG", "ASSISTANTE SGA", "Resp SECURITÉ", "AGENT SECURITÉ",
                                              "CHAUFFEUR", "PLANTON"])))
        self.title.grid(row=0, column=0)
        self.title_combox.grid(row=1, column=0, ipadx=50)

        # Create department label and combobox
        self.department_label = ttk.Label(self.user_info_frame, text="DEPARTEMENT:", background='lightgrey',
                                          underline=0)
        self.department_combobox = ttk.Combobox(self.user_info_frame,
                                                values=(sorted(["CABINET", "SECRETARIAT", "DAFC",
                                                                "PARTENARIAT",
                                                                "CELLULE NUTRITION", "NEXUS",
                                                                "COORDINATION REG",
                                                                "DMRC", "DSEC", "DPEP", "GNN-SECURITY",
                                                                "AUXILIAIRES"])))
        self.department_label.grid(row=0, column=1)
        self.department_combobox.grid(row=1, column=1, ipadx=20)

        # The place/SIEGE combobox
        self.place_label = ttk.Label(self.user_info_frame, text='LIEU:', background='lightgrey', underline=0)
        self.place_label.grid(row=0, column=2)
        self.site_name_list = ['SIEGE', 'ANNEXE 1', 'ANNEXE 2']
        self.place_combobox = ttk.Combobox(self.user_info_frame, values=self.site_name_list,
                                           validate='focus')
        self.place_combobox.grid(row=1, column=2)

        # Create break time start and end labels
        self.break_start_label = ttk.Label(self.user_info_frame, text="DEBUT PAUSE (HH:MM):", background="orange",
                                           underline=0)
        self.break_end_label = ttk.Label(self.user_info_frame, text="RETOUR PAUSE (HH:MM):", background="orange",
                                         underline=0)
        self.break_start_label.grid(row=2, column=1)
        self.break_end_label.grid(row=2, column=2)

        # Create start and end time labels
        self.time_start_label = ttk.Label(self.user_info_frame, text="HEURE ENTREE (HH:MM):",
                                          background="lightgreen", underline=0)
        self.time_end_label = ttk.Label(self.user_info_frame, text="DESCENTE (HH:MM):", background="red",
                                        underline=0)
        self.time_start_label.grid(row=2, column=0)
        self.time_end_label.grid(row=2, column=3)

        # Create Entry widgets for arrival and departure time entries
        self.time_start_entry = ttk.Entry(self.user_info_frame)
        self.time_end_entry = ttk.Entry(self.user_info_frame)
        self.time_start_entry.grid(row=3, column=0)
        self.time_end_entry.grid(row=3, column=3)

        # Create entry widget for break starts and break ends time entries
        self.break_start_entry = ttk.Entry(self.user_info_frame)
        self.break_end_entry = ttk.Entry(self.user_info_frame)
        self.break_start_entry.grid(row=3, column=1)
        self.break_end_entry.grid(row=3, column=2)

        # Week days combobox entry
        self.week_label = ttk.Label(self.user_info_frame, text="JOUR DE SEMAINE:", underline=0, background="lightgrey")
        combobox = ttk.Combobox(self.user_info_frame,
                                values=["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"])
        self.week_combobox = combobox
        self.week_combobox.grid(row=1, column=3)
        self.week_label.grid(row=0, column=3)

        for widget in self.user_info_frame.winfo_children():
            widget.grid_configure(padx=20, pady=5, sticky="news")

        # # Create the message to display
        ## create labelframe to display the message within it

        self.msg_labelframe = ttk.LabelFrame(self.frame, text='HC3N', underline=0, labelanchor='n')
        self.msg_labelframe.grid(row=4, column=1, sticky="news", padx=10, pady=10)
        self.ourmessage = "HAUT COMMISSARIAT A L'INITIATIVE 3N"
        self.display_message = tk.Message(self.msg_labelframe, text=self.ourmessage, font='italic')
        self.display_message.config(justify="center", bd=5, highlightthickness=0, highlightcolor='blue', relief='sunken')
        self.display_message.grid(row=0, column=0, ipadx=50)

        ### Display time
        def display_time():
            """
            Print a clock time
            :return: A numeric clock
            """
            string = strftime("%d-%m-%Y %H:%M:%S")
            self.label_text = ttk.Label(self.msg_labelframe, font=("arial nova", 14, 'italic'), foreground='black',
                                        background='lightgrey', relief='sunken')
            self.label_text.config(text=string)
            self.label_text.grid(row=0, column=3, sticky='news', ipadx=20, pady=10, padx=10)
            self.label_text.after(1000, display_time)

        display_time()

        ############################################ Configure new_frame LabelFrame ################################
        self.new_frame = ttk.LabelFrame(self.frame, text="SORTIE & ENTREE SUPPLEMENTAIRE", underline=0)
        self.new_frame.grid(row=3, column=1, sticky='news', padx=15, pady=15)

        self.new_label = ttk.Label(self.new_frame, text="HEURE DEPART (HH:MM):", underline=0, background="lightgreen")
        self.second_new_label = ttk.Label(self.new_frame, text="HEURE RETOUR (HH:MM):", underline=0,
                                          background="red")

        self.new_label.grid(row=0, column=0)
        self.second_new_label.grid(row=0, column=1)

        self.new_entry = ttk.Entry(self.new_frame)
        self.new_exit = ttk.Entry(self.new_frame)

        self.new_entry.grid(row=1, column=0)
        self.new_exit.grid(row=1, column=1)

        self.first_btn_check_var = tk.BooleanVar(self.new_frame, value=False)
        self.second_btn_check_var = tk.BooleanVar(self.new_frame, value=False)

        self.first_btn_checkbutton = ttk.Checkbutton(self.new_frame, text="Sortie: Cadre du travail / Autorisée",
                                                     variable=self.first_btn_check_var, onvalue=True, offvalue=False,
                                                     underline=0)
        self.btn_second_checkbutton = ttk.Checkbutton(self.new_frame, text="Sortie: Hors Cadre du travail",
                                                      variable=self.second_btn_check_var, onvalue=True, offvalue=False,
                                                      underline=0)

        self.first_btn_checkbutton.grid(row=0, column=2)
        self.btn_second_checkbutton.grid(row=2, column=2)

        self.personal_label = ttk.Label(self.new_frame, text="HEURE DEPART (HH:MM):", background="lightgreen",
                                        underline=0)
        self.personal_second_label = ttk.Label(self.new_frame, text="HEURE RETOUR (HH:MM):", background="red",
                                               underline=0)

        self.personal_entry = ttk.Entry(self.new_frame)
        self.personal_exit = ttk.Entry(self.new_frame)

        # Display
        self.personal_label.grid(row=2, column=0)
        self.personal_entry.grid(row=3, column=0)
        self.personal_second_label.grid(row=2, column=1)
        self.personal_exit.grid(row=3, column=1)

        for widget in self.new_frame.winfo_children():
            widget.grid_configure(padx=20, pady=5, sticky="news")

        ############################################### Second LabelFrame ###################################################

        # Create the exit/entry, label and checkButtons from HQ to annexe1-2

        self.exit_entry_frame = ttk.LabelFrame(self.frame, text="EQUIPE SIEGE vers (ANNEXE-1 et ANNEXE-2)", underline=0)
        self.exit_entry_frame.grid(row=1, column=0, sticky="news", padx=15, pady=10)

        self.exit_entry_status_var_1 = tk.BooleanVar(self.exit_entry_frame, value=False)
        self.exit_entry_status_check_1 = ttk.Checkbutton(self.exit_entry_frame, text="Verifier Presence ANNEXE-1",
                                                         variable=self.exit_entry_status_var_1, onvalue=True,
                                                         offvalue=False, underline=9)
        self.exit_entry_status_check_1.grid(row=0, column=0)

        # Create Entry widgets for exit/entry from Site
        self.exit_entry_label_entry = ttk.Label(self.exit_entry_frame, text="HEURE ENTREE (HH:MM):",
                                                background="lightgreen", underline=0)
        self.exit_entry_label_exit = ttk.Label(self.exit_entry_frame, text="SORTIE (HH:MM):",
                                               background="red", underline=0)
        self.exit_entry_label_entry.grid(row=1, column=0)
        self.exit_entry_label_exit.grid(row=1, column=1)

        self.site_entry = ttk.Entry(self.exit_entry_frame)
        self.site_exit = ttk.Entry(self.exit_entry_frame)
        self.site_entry.grid(row=2, column=0)
        self.site_exit.grid(row=2, column=1)

        ############# second entry/exit Entry widget #############

        self.exit_entry_status_var_2 = tk.BooleanVar(self.exit_entry_frame, value=False)
        self.exit_entry_status_check_2 = ttk.Checkbutton(self.exit_entry_frame, text="Verifier Presence Annexe-2",
                                                         variable=self.exit_entry_status_var_2, onvalue=True,
                                                         offvalue=False, underline=9)
        self.exit_entry_status_check_2.grid(row=0, column=3)

        # Create Entry Exit Status from HQ to annexe1-2
        self.exit_entry_label_entry_01 = ttk.Label(self.exit_entry_frame, text="HEURE ENTREE (HH:MM):",
                                                   background="lightgreen", underline=0)
        self.exit_entry_label_exit_01 = ttk.Label(self.exit_entry_frame, text="SORTIE (HH:MM):",
                                                  background="red", underline=0)
        self.exit_entry_label_entry_01.grid(row=1, column=3)
        self.exit_entry_label_exit_01.grid(row=1, column=4)

        self.site_entry_01 = ttk.Entry(self.exit_entry_frame)
        self.site_exit_01 = ttk.Entry(self.exit_entry_frame)
        self.site_entry_01.grid(row=2, column=3)
        self.site_exit_01.grid(row=2, column=4)

        for widget in self.exit_entry_frame.winfo_children():
            widget.grid_configure(padx=20, pady=6, sticky="news")

        ###################################### Configure Third LabelFrame ################################

        # Create the labelFrame annex 1 to hq and annexe 2
        ################# function_set check, entries and labels
        self.verify_frame = ttk.LabelFrame(self.frame, text="EQUIPE ANNEXE-1 vers (SIEGE et ANNEXE-2)", underline=0)
        self.verify_frame.grid(row=2, column=0, padx=20, pady=15, sticky="news")
        self.presence_check_var = tk.BooleanVar(self.verify_frame, value=False)
        self.presence_check = ttk.Checkbutton(self.verify_frame, text="Verifier Presence SIEGE",
                                              variable=self.presence_check_var,
                                              onvalue=True, offvalue=False, underline=9)
        self.presence_check.grid(row=0, column=0)

        # Verifier presence Entry and Exit / label and entry widgets
        self.annexe_entry_label = ttk.Label(self.verify_frame, text="HEURE ENTREE (HH:MM):",
                                            background="lightgreen",
                                            underline=0)
        self.annexe_exit_label = ttk.Label(self.verify_frame, text="SORTIE (HH:MM):", background="red",
                                           underline=0)
        self.annexe_entry_label.grid(row=1, column=0)
        self.annexe_exit_label.grid(row=1, column=1)

        self.annexe_entry = ttk.Entry(self.verify_frame)
        self.annexe_exit = ttk.Entry(self.verify_frame)
        self.annexe_entry.grid(row=2, column=0)
        self.annexe_exit.grid(row=2, column=1)

        ####################

        # second check variable, labels and Entries widgets
        self.annexe_to_annexe_var = tk.BooleanVar(self.verify_frame, value=False)
        self.annexe_to_annexe = ttk.Checkbutton(self.verify_frame, text="Verifier Presence ANNEXE-2",
                                                variable=self.annexe_to_annexe_var, onvalue=True, offvalue=False,
                                                underline=9)
        self.annexe_to_annexe.grid(row=0, column=3)

        self.check_annexe_entry_label = ttk.Label(self.verify_frame, text="HEURE ENTREE (HH:MM):",
                                                  background="lightgreen", underline=0)
        self.check_annexe_exit_label = ttk.Label(self.verify_frame, text="SORTIE (HH:MM):", background="red",
                                                 underline=0)

        self.check_annexe_entry_label.grid(row=1, column=3)
        self.check_annexe_exit_label.grid(row=1, column=4)

        # Create entry widget
        self.annexe_entry_01 = ttk.Entry(self.verify_frame)
        self.annexe_exit_01 = ttk.Entry(self.verify_frame)
        self.annexe_entry_01.grid(row=2, column=3)
        self.annexe_exit_01.grid(row=2, column=4)

        for widget in self.verify_frame.winfo_children():
            widget.grid_configure(padx=20, pady=5, sticky="news")

        ############################################ Configure Fourth LabelFrame ################################

        self.third_frame = ttk.LabelFrame(self.frame, text="EQUIPE ANNEXE-2 vers (SIEGE-ANNEXE-1)", underline=0)
        self.third_frame.grid(row=3, column=0, sticky="news", padx=20, pady=10)

        self.verification_button_var = tk.BooleanVar(self.third_frame, value=False)
        self.verification_button = ttk.Checkbutton(self.third_frame, text="Verifier Presence SIEGE",
                                                   variable=self.verification_button_var,
                                                   onvalue=True, offvalue=False, underline=9)
        self.verification_button.grid(row=0, column=0)

        self.value_entry_label = ttk.Label(self.third_frame, text="HEURE ENTREE (HH:MM):",
                                           background="lightgreen", underline=0)
        self.value_exit_label = ttk.Label(self.third_frame, text="SORTIE (HH:MM):", background="red", underline=0)
        self.value_entry_widget = ttk.Entry(self.third_frame)
        self.value_exit_widget = ttk.Entry(self.third_frame)

        self.value_entry_label.grid(row=1, column=0)
        self.value_exit_label.grid(row=1, column=1)
        self.value_entry_widget.grid(row=2, column=0)
        self.value_exit_widget.grid(row=2, column=1)

        ###### second verification check button, entries

        self.second_verification_check_var = tk.BooleanVar(self.third_frame, value=False)
        self.second_verification_check_button = ttk.Checkbutton(self.third_frame, text="Verifier Presence ANNEXE-1",
                                                                variable=self.second_verification_check_var,
                                                                onvalue=True, offvalue=False, underline=9)
        self.second_verification_check_button.grid(row=0, column=3)

        self.second_ver_label1 = ttk.Label(self.third_frame, text="HEURE ENTREE (HH:MM):",
                                           background="lightgreen", underline=0)
        self.second_ver_label2 = ttk.Label(self.third_frame, text="SORTIE (HH:MM):", background="red",
                                           underline=0)
        self.second_annexe_entry = ttk.Entry(self.third_frame)
        self.second_annexe_exit = ttk.Entry(self.third_frame)
        self.second_ver_label1.grid(row=1, column=3)
        self.second_ver_label2.grid(row=1, column=4)
        self.second_annexe_entry.grid(row=2, column=3)
        self.second_annexe_exit.grid(row=2, column=4)

        for widget in self.third_frame.winfo_children():
            widget.grid_configure(padx=20, pady=5, sticky="news")

        ############################################ Configure break, total autofill, auto-delete LabelFrame ################################

        self.registration_frame = ttk.LabelFrame(self.frame, text='PAUSE - SAUVEGARDE - CALCUL TEMPS TOTAL',
                                                 underline=0)
        self.registration_frame.grid(row=4, column=0, sticky='news', padx=20, pady=10)

        self.break_check_button_var = tk.BooleanVar(self.registration_frame, value=False)
        self.break_check_button = ttk.Checkbutton(self.registration_frame, text="Pause entre 13:30 PM et 14:15 PM",
                                                  variable=self.break_check_button_var, onvalue=True, offvalue=False,
                                                  underline=0)
        self.break_check_button.grid(row=2, column=0)

        # result view label
        self.result_view = ttk.Label(self.registration_frame, text="Le Temps Total est --> :", underline=3,
                                     background="lightgrey")
        self.result_view.grid(row=2, column=1, sticky='news')
        self.result_label = ttk.Label(self.registration_frame, background='lightgreen')
        self.result_label.grid(row=2, column=2)

        self.onsite_check_var = tk.BooleanVar(self.registration_frame, value=False)
        self.onsite_confirm_presence = ttk.Checkbutton(self.registration_frame, text="Absence / Congé",
                                                       variable=self.onsite_check_var, onvalue=True, offvalue=False,
                                                       underline=9)
        self.onsite_confirm_presence.grid(row=2, column=5)

        for widget in self.registration_frame.winfo_children():
            widget.grid_configure(padx=10, pady=15, sticky="news")

        ################################################################

        #### New label for display the duration time at a site
        self.stay_time_labelframe = ttk.Labelframe(self.frame,
                                                   text="AFFICHAGE DUREE: 'PAUSE - ENTREE-SORTIE - DEPLACEMENT EQUIPE'",
                                                   underline=0)
        self.stay_time_labelframe.grid(row=1, column=1, sticky='news', padx=20, pady=10)

        # self.duration_tion_label = ttk.Label(self.stay_time_labelframe)
        # self.duration_tion_label.grid(row=0, column=0, sticky='news', padx=20, pady=10)
        self.display_duration_label = ttk.Label(self.stay_time_labelframe, background='lightgreen')
        self.display_duration_label.grid(row=0, column=1, sticky='news', padx=20, pady=10)

        self.standard_duration_label = ttk.Label(self.stay_time_labelframe, text="TEMPS STANDARD A PASSER AU BUREAU:", background='lightgrey', underline=0)
        self.standard_duration_label.grid(row=2, column=0)

        self.standard_duration_entry = ttk.Entry(self.stay_time_labelframe)
        self.standard_duration_entry.grid(row=2, column=1)

        for widget in self.stay_time_labelframe.winfo_children():
            widget.grid_configure(padx=10, pady=15, sticky="news")

        #############################

        """
            # Create Buttons
        """

        # create calculate button
        self.calculate_button = ttk.Button(self.registration_frame, text="Calculer Temps Total",
                                           command=self.calculate_total_time,
                                           underline=0)
        self.calculate_button.grid(row=3, column=2, sticky='news', padx=10)
        ############ Create save_to_excel Button
        self.entry_button = ttk.Button(self.registration_frame, text='Sauvegarder', command=self.save_to_excel,
                                       underline=0)
        self.entry_button.grid(row=3, column=0, sticky='news', padx=10)

        ################ add a clear button
        self.clear_button = ttk.Button(self.registration_frame, text='Effacer Champs', command=self.clear, underline=0)
        self.clear_button.grid(row=3, column=1, padx=10)

        ################# Auto fill Button
        self.auto_fill_button = ttk.Button(self.registration_frame, text='Remplissage Automatique', command=self.fill_entries,
                                           underline=0)
        self.auto_fill_button.grid(row=3, column=4, padx=10)

        ##### Duration time button
        self.duration_time_button = ttk.Button(self.stay_time_labelframe, text="Afficher Durée",
                                               command=self.display_duration,
                                               underline=0, )
        self.duration_time_button.grid(row=0, column=0, sticky='news', padx=10, pady=10)

        #################################### Configure the Menu #################################

        # Create Menu menubar
        self.menu_ = tk.Menu(self.frame, tearoff=0)
        self.menu_bar = tk.Menu(self.menu_, tearoff=0)
        self.menu_bar.add_command(label="A propos HC3N", command=on_click)
        self.menu_bar.add_separator()
        self.menu_bar.add_command(label="Quitter", accelerator='Alt+F4', command=self.exit_01, underline=0)

        self.menu_.add_cascade(label="Menu", menu=self.menu_bar, hidemargin=True)
        self.window.config(menu=self.menu_)

        #
        """
            # Everything about functions
        """

    ##### Calendar function

    # Exit function
    def exit_01(self):
        """
        Exit function
        :return: Destroys window and exit the program
        """
        if messagebox.askokcancel(title='Quitter', message='Voulez-vous quitter ?'):
            self.window.destroy()

        # Clear function

    def clear(self):
        """
        Delete all values from entries
        :return: clear all values from entries
        """
        self.first_last_name_entry.delete(0, END)
        self.place_combobox.delete(0, END)
        self.title_combox.delete(0, END)
        self.department_combobox.delete(0, END)
        self.week_combobox.delete(0, END)
        self.time_start_entry.delete(0, END)
        self.time_end_entry.delete(0, END)
        self.break_start_entry.delete(0, END)
        self.break_end_entry.delete(0, END)
        self.annexe_entry.delete(0, END)
        self.annexe_entry_01.delete(0, END)
        self.site_exit.delete(0, END)
        self.site_exit_01.delete(0, END)
        self.site_entry.delete(0, END)
        self.site_entry_01.delete(0, END)
        self.annexe_exit_01.delete(0, END)
        self.annexe_exit.delete(0, END)
        self.value_entry_widget.delete(0, END)
        self.value_exit_widget.delete(0, END)
        self.second_annexe_entry.delete(0, END)
        self.second_annexe_exit.delete(0, END)
        self.observation_list_combobox.delete(0, END)
        self.date_entry.delete(0, END)
        self.new_entry.delete(0, END)
        self.new_exit.delete(0, END)
        self.personal_exit.delete(0, END)
        self.personal_entry.delete(0, END)
        self.standard_duration_entry.delete(0, END)

        ################################

    def fill_entries(self):
        """
        fill out all entries with default value "00:00"
        :return:
        """
        value = "00:00"  # Predefined value to fill the entries
        second = "08:00"  # Predefined second value to fill the entries
        third = "17:30"  # Predefined third value to fill the entries
        fourth = "13:30"  # Predefined fourth value to fill the entries
        fifth = "14:15"  # Predefined fifth value to fill the entries
        sixth = "00:01"  # Predefined 6th value to fill the entries
        seventh = "13:00"  # Predefined seventh value to fill the entries
        eighth = str(datetime.timedelta(hours=8, minutes=45))  # Predefined eighth to fill the entries
        nineth = str(datetime.timedelta(hours=5, minutes=0))  # Predefined nineth to fill the entries

        if self.break_check_button_var.get():
            self.break_start_entry.insert(0, fourth)
            self.break_end_entry.insert(0, fifth)
        else:
            self.break_start_entry.insert(0, value)
            self.break_end_entry.insert(0, value)

        if self.onsite_check_var.get():
            self.time_end_entry.insert(0, sixth)
            self.time_start_entry.insert(0, value)
        elif self.week_combobox.get() == 'Vendredi':
            self.time_start_entry.insert(0, second)
            self.time_end_entry.insert(0, seventh)
        else:
            self.time_start_entry.insert(0, second)
            self.time_end_entry.insert(0, third)

        if self.week_combobox.get() == 'Vendredi':
            self.standard_duration_entry.insert(0, nineth)
        else:
            self.standard_duration_entry.insert(0, eighth)

        self.annexe_entry.insert(0, value)
        self.annexe_entry_01.insert(0, value)
        self.site_exit.insert(0, value)
        self.site_exit_01.insert(0, value)
        self.site_entry.insert(0, value)
        self.site_entry_01.insert(0, value)
        self.annexe_exit_01.insert(0, value)
        self.annexe_exit.insert(0, value)
        self.value_entry_widget.insert(0, value)
        self.value_exit_widget.insert(0, value)
        self.second_annexe_entry.insert(0, value)
        self.second_annexe_exit.insert(0, value)
        self.new_entry.insert(0, value)
        self.new_exit.insert(0, value)
        self.personal_entry.insert(0, value)
        self.personal_exit.insert(0, value)

    ################################################################

    #  Calculate function

    def calculate_total_time(self):
        """
        A function that calculates the total time of multiple entries as time formatted
        :return: The total time
        """
        start_time_str = self.time_start_entry.get()
        end_time_str = self.time_end_entry.get()
        break_start_time_str = self.break_start_entry.get()  # fixed break start time
        break_end_time_str = self.break_end_entry.get()  # fixed break end time
        break_taken = self.break_check_button_var.get()
        # standard_time_str = self.standard_duration_entry.get()

        start_time = datetime.datetime.strptime(start_time_str, "%H:%M").time()
        end_time = datetime.datetime.strptime(end_time_str, "%H:%M").time()
        break_start_time = datetime.datetime.strptime(break_start_time_str, "%H:%M").time()
        break_end_time = datetime.datetime.strptime(break_end_time_str, "%H:%M").time()
        # standard_time = datetime.datetime.strptime(standard_time_str, "%H:%M:%S").time()

        # TEAM HQ TO (ANNEXE"1-2")
        hq_to_annexe1_entry_str = self.site_entry.get()
        hq_to_annexe1_exit_str = self.site_exit.get()
        hq_to_annexe2_entry_str = self.site_entry_01.get()
        hq_to_annexe2_exit_str = self.site_exit_01.get()

        # Conditional Verification Team HQ
        hq_to_annexe1_entry = datetime.datetime.strptime(hq_to_annexe1_entry_str, "%H:%M").time()
        hq_to_annexe1_exit = datetime.datetime.strptime(hq_to_annexe1_exit_str, "%H:%M").time()
        hq_to_annexe2_entry = datetime.datetime.strptime(hq_to_annexe2_entry_str, "%H:%M").time()
        hq_to_annexe2_exit = datetime.datetime.strptime(hq_to_annexe2_exit_str, "%H:%M").time()
        hq_visit_to_annexe1_check = self.exit_entry_status_var_1.get()
        hq_visit_to_annexe2_check = self.exit_entry_status_var_2.get()

        # Team ANNEXE 1 TO (HQ-ANNEXE 2)
        annexe1_to_hq_entry_str = self.annexe_entry.get()
        annexe1_to_hq_exit_str = self.annexe_exit.get()
        annexe1_to_annexe2_entry_str = self.annexe_entry_01.get()
        annexe1_to_annexe2_exit_str = self.annexe_exit_01.get()

        # Conditional Verification Team annexe1
        annexe1_to_hq_entry = datetime.datetime.strptime(annexe1_to_hq_entry_str, "%H:%M").time()
        annexe1_to_hq_exit = datetime.datetime.strptime(annexe1_to_hq_exit_str, "%H:%M").time()
        annexe1_to_annexe2_entry = datetime.datetime.strptime(annexe1_to_annexe2_entry_str, "%H:%M").time()
        annexe1_to_annexe2_exit = datetime.datetime.strptime(annexe1_to_annexe2_exit_str, "%H:%M").time()
        annexe1_visit_to_hq_check = self.presence_check_var.get()
        annexe1_visit_to_annexe2_check = self.annexe_to_annexe_var.get()

        # TEAM ANNEXE 2 TO (HQ-ANNEXE 1)
        annexe2_to_hq_entry_str = self.value_entry_widget.get()
        annexe2_to_hq_exit_str = self.value_exit_widget.get()
        annexe2_to_annexe1_entry_str = self.second_annexe_entry.get()
        annexe2_to_annexe1_exit_str = self.second_annexe_exit.get()

        # Conditional Verification Team Annexe2
        annexe2_to_hq_entry = datetime.datetime.strptime(annexe2_to_hq_entry_str, "%H:%M").time()
        annexe2_to_hq_exit = datetime.datetime.strptime(annexe2_to_hq_exit_str, "%H:%M").time()
        annexe2_to_annexe1_entry = datetime.datetime.strptime(annexe2_to_annexe1_entry_str, "%H:%M").time()
        annexe2_to_annexe1_exit = datetime.datetime.strptime(annexe2_to_annexe1_exit_str, "%H:%M").time()
        annexe2_visit_to_hq_check = self.verification_button_var.get()
        annexe2_visit_to_annexe1_check = self.second_verification_check_var.get()

        ### Complementary Entry / Exit
        new_entry_str = self.new_entry.get()
        second_exit_str = self.new_exit.get()
        personal_entry_str = self.personal_entry.get()
        personal_exit_str = self.personal_exit.get()

        #
        new_entry = datetime.datetime.strptime(new_entry_str, "%H:%M").time()
        new_exit = datetime.datetime.strptime(second_exit_str, "%H:%M").time()
        personal_entry = datetime.datetime.strptime(personal_entry_str, "%H:%M").time()
        personal_exit = datetime.datetime.strptime(personal_exit_str, "%H:%M").time()
        work_case_exit = self.first_btn_check_var.get()
        personal_case_exit = self.second_btn_check_var.get()

        # Not present at office
        not_present = self.onsite_check_var.get()

        # Usage of whole conditionals variables
        try:
            # total_time = datetime.timedelta()  # Initialize total_time to zero
            if not_present:
                messagebox.showwarning(title='Abscence Signalée!', message="L'employé(e) a été absent(e)!")
            if not break_taken:
                if not (
                        hq_visit_to_annexe1_check or hq_visit_to_annexe2_check or annexe1_visit_to_hq_check or annexe1_visit_to_annexe2_check or annexe2_visit_to_hq_check or annexe2_visit_to_annexe1_check or work_case_exit or personal_case_exit):
                    messagebox.showwarning(title="Alerte",
                                           message="1.L'employé(e) n'a pas prit de pause ! \n\n 2.Aucun deplacement effectué vers: \n\n 'SIEGE;  ANNEXE-1;  ANNEXE-2'; \n\n 3.Aucune Sortie Signalée!")

                total_time = datetime.datetime.combine(datetime.date.today(),
                                                       end_time) - datetime.datetime.combine(
                    datetime.date.today(), start_time)

                if end_time <= start_time:
                    total_time = datetime.timedelta(hours=0, minutes=0, seconds=0)

                if hq_visit_to_annexe1_check:
                    messagebox.showwarning(title='Alerte',
                                           message=f"L'employé(e) du SIEGE s'est rendu a l'ANNEXE-1 entre: \n\n{hq_to_annexe1_entry} et {hq_to_annexe1_exit}")
                    # total_time = datetime.timedelta(hours=17, minutes=30)
                    if hq_to_annexe1_exit > end_time:
                        total_time = datetime.datetime.combine(datetime.date.today(),
                                                               end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time) + datetime.datetime.combine(datetime.date.today(),
                                                                    hq_to_annexe1_exit) - datetime.datetime.combine(
                            datetime.date.today(), hq_to_annexe1_entry)

                    elif end_time > hq_to_annexe1_exit:
                        total_time = datetime.datetime.combine(datetime.date.today(),
                                                               end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)

                if hq_visit_to_annexe2_check:
                    messagebox.showwarning(title="Alerte",
                                           message=f"L'employé(e) du SIEGE s'est rendu a l'ANNEXE-2 entre: \n\n{hq_to_annexe2_entry} et {hq_to_annexe2_exit}")
                    # total_time = datetime.timedelta(hours=17, minutes=30)
                    if hq_to_annexe2_exit >= end_time:
                        total_time = datetime.datetime.combine(datetime.date.today(),
                                                               end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time) + datetime.datetime.combine(datetime.date.today(),
                                                                    hq_to_annexe2_exit) - datetime.datetime.combine(
                            datetime.date.today(), hq_to_annexe2_entry)

                    elif end_time > hq_to_annexe2_exit:
                        total_time = datetime.datetime.combine(datetime.date.today(),
                                                               end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)

                if annexe1_visit_to_hq_check:
                    messagebox.showwarning(title="Alerte",
                                           message=f"L'employé(e) ANNEXE-1 s'est rendu au SIEGE entre: \n\n{annexe1_to_hq_entry} et {annexe1_to_hq_exit}")
                    # total_time = datetime.timedelta(hours=17, minutes=30)
                    if annexe1_to_hq_exit > end_time:
                        total_time = datetime.datetime.combine(datetime.date.today(),
                                                               end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time) + datetime.datetime.combine(datetime.date.today(),
                                                                    annexe1_to_hq_exit) - datetime.datetime.combine(
                            datetime.date.today(), annexe1_to_hq_entry)

                    elif end_time > annexe1_to_hq_exit:
                        total_time = datetime.datetime.combine(datetime.date.today(),
                                                               end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)

                if annexe1_visit_to_annexe2_check:
                    messagebox.showwarning(title="Alerte",
                                           message=f"L'employé(e) ANNEXE-1 s'est rendu a l'ANNEXE-2 entre: \n\n{annexe1_to_annexe2_entry} et {annexe1_to_annexe2_exit}")
                    # total_time = datetime.timedelta(hours=17, minutes=30)
                    if annexe1_to_annexe2_exit > end_time:
                        total_time = datetime.datetime.combine(datetime.date.today(),
                                                               end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time) + datetime.datetime.combine(datetime.date.today(),
                                                                    annexe1_to_annexe2_exit) - datetime.datetime.combine(
                            datetime.date.today(), annexe1_to_annexe2_entry)

                    elif end_time > annexe1_to_annexe2_exit:
                        total_time = datetime.datetime.combine(datetime.date.today(),
                                                               end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)

                if annexe2_visit_to_hq_check:
                    messagebox.showwarning(title="Alerte",
                                           message=f"L'employé(e) ANNEXE-2 s'est rendu au SIEGE entre: \n\n{annexe2_to_hq_entry} et {annexe2_to_hq_exit}")
                    # total_time = datetime.timedelta(hours=17, minutes=30)
                    if annexe2_to_hq_exit > end_time:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) + (datetime.datetime.combine(datetime.date.today(),
                                                                      annexe2_to_hq_exit) -
                                            datetime.datetime.combine(datetime.date.today(), annexe2_to_hq_entry))

                    elif end_time > annexe2_to_hq_exit:
                        total_time = datetime.datetime.combine(datetime.date.today(),
                                                               end_time) - datetime.datetime.combine(
                            datetime.date.today(), start_time)

                if annexe2_visit_to_annexe1_check:
                    messagebox.showwarning(title="Alerte",
                                           message=f"L'employé(e) ANNEXE-2 s'est rendu a l'ANNEXE-1 entre: \n\n{annexe2_to_annexe1_entry} et {annexe2_to_annexe1_exit}")
                    # total_time = datetime.timedelta(hours=17, minutes=30)
                    if annexe2_to_annexe1_exit > end_time:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) + (datetime.datetime.combine(datetime.date.today(),
                                                                      annexe2_to_annexe1_exit) -
                                            datetime.datetime.combine(datetime.date.today(), annexe2_to_annexe1_entry))

                    elif end_time > annexe2_to_annexe1_exit:
                        total_time = datetime.datetime.combine(datetime.date.today(),
                                                               end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)

                if work_case_exit:
                    messagebox.showwarning(title="Alerte",
                                           message=f"Sortie Signalée: Cadre du Travail / Autorisée entre: \n\n{new_entry} et {new_exit}")
                    # total_time = datetime.timedelta(hours=17, minutes=30)
                    if new_exit > end_time:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) + (datetime.datetime.combine(datetime.date.today(),
                                                                      new_exit) -
                                            datetime.datetime.combine(datetime.date.today(), new_entry))

                    elif end_time > new_exit:
                        total_time = datetime.datetime.combine(datetime.date.today(),
                                                               end_time) - datetime.datetime.combine(
                            datetime.date.today(), start_time)

                if personal_case_exit:
                    messagebox.showwarning(title="Alerte",
                                           message=f"Sortie Signalée: Hors Cadre du Travail entre: \n\n{personal_entry} et {personal_exit}")
                    total_time -= datetime.datetime.combine(datetime.date.today(),
                                                            personal_exit) - datetime.datetime.combine(
                        datetime.date.today(),
                        personal_entry)

            else:
                # total_time = datetime.timedelta()
                messagebox.showinfo(title="Information",
                                    message=f"L'Employé(e) a prit une pause entre: \n\n{break_start_time} et {break_end_time}")
                total_time = (datetime.datetime.combine(datetime.date.today(),
                                                        end_time) - datetime.datetime.combine(
                    datetime.date.today(), start_time)) - (
                                     datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                     datetime.datetime.combine(datetime.date.today(), break_start_time))

                if not (hq_visit_to_annexe1_check or hq_visit_to_annexe2_check or annexe1_visit_to_hq_check or
                        annexe1_visit_to_annexe2_check or annexe2_visit_to_hq_check or annexe2_visit_to_annexe1_check):
                    if not (work_case_exit or personal_case_exit):
                        messagebox.showwarning(title="Alerte",
                                               message="1.Aucun deplacement effectuee vers: \n\n 'SIEGE;  ANNEXE-1;  ANNEXE-2'; \n\n 2.Aucune Sortie Signalée!")

                if break_start_time > break_end_time:
                    messagebox.showerror(title="Erreur", message="Temps de pause Incorrect.")

                # if start_time < break_start_time and end_time >= break_end_time:

                elif break_start_time == break_end_time:
                    messagebox.showerror(title="Erreur",
                                         message="Heure Debut pause est supérieure ou égal a l'heure de Retour de pause!")
                    total_time = datetime.datetime.combine(datetime.date.today(),
                                                           end_time) - datetime.datetime.combine(
                        datetime.date.today(), start_time)

                if hq_visit_to_annexe1_check:
                    messagebox.showwarning(title='Alerte',
                                           message=f"L'employé(e) du SIEGE s'est rendu a l'ANNEXE-1 entre: \n\n{hq_to_annexe1_entry} et {hq_to_annexe1_exit}")
                    # total_time = datetime.timedelta(hours=17, minutes=30)
                    if hq_to_annexe1_exit > end_time:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) - (
                                             datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                             datetime.datetime.combine(datetime.date.today(),
                                                                       break_start_time)) + (
                                             datetime.datetime.combine(datetime.date.today(),
                                                                       hq_to_annexe1_exit) -
                                             datetime.datetime.combine(datetime.date.today(), hq_to_annexe1_entry))

                    elif end_time > hq_to_annexe1_exit:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) - (
                                             datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                             datetime.datetime.combine(datetime.date.today(), break_start_time))

                if hq_visit_to_annexe2_check:
                    messagebox.showwarning(title="Alerte",
                                           message=f"L'employé(e) du SIEGE s'est rendu a l'ANNEXE-2 entre: \n\n{hq_to_annexe2_entry} et {hq_to_annexe2_exit}")
                    # total_time = datetime.timedelta(hours=17, minutes=30)
                    if hq_to_annexe2_exit >= end_time:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) - (
                                             datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                             datetime.datetime.combine(datetime.date.today(),
                                                                       break_start_time)) + datetime.datetime.combine(
                            datetime.date.today(),
                            hq_to_annexe2_exit) - datetime.datetime.combine(
                            datetime.date.today(), hq_to_annexe2_entry)

                    elif end_time > hq_to_annexe2_exit:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) - (
                                             datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                             datetime.datetime.combine(datetime.date.today(), break_start_time))

                if annexe1_visit_to_hq_check:
                    messagebox.showwarning(title="Alerte",
                                           message=f"L'employé(e) ANNEXE-1 s'est rendu au SIEGE entre: \n\n{annexe1_to_hq_entry} et {annexe1_to_hq_exit}")
                    # total_time = datetime.timedelta(hours=17, minutes=30)
                    if annexe1_to_hq_exit > end_time:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) - (
                                             datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                             datetime.datetime.combine(datetime.date.today(),
                                                                       break_start_time)) + datetime.datetime.combine(
                            datetime.date.today(),
                            annexe1_to_hq_exit) - datetime.datetime.combine(
                            datetime.date.today(), annexe1_to_hq_entry)

                    elif end_time > annexe1_to_hq_exit:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) - (
                                             datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                             datetime.datetime.combine(datetime.date.today(), break_start_time))

                if annexe1_visit_to_annexe2_check:
                    messagebox.showwarning(title="Alerte",
                                           message=f"L'employé(e) ANNEXE-1 s'est rendu a l'ANNEXE-2 entre: \n\n{annexe1_to_annexe2_entry} et {annexe1_to_annexe2_exit}")
                    # total_time = datetime.timedelta(hours=17, minutes=30)
                    if annexe1_to_annexe2_exit > end_time:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) - (
                                             datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                             datetime.datetime.combine(datetime.date.today(),
                                                                       break_start_time)) + datetime.datetime.combine(
                            datetime.date.today(),
                            annexe1_to_annexe2_exit) - datetime.datetime.combine(
                            datetime.date.today(), annexe1_to_annexe2_entry)

                    elif end_time > annexe1_to_annexe2_exit:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) - (
                                             datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                             datetime.datetime.combine(datetime.date.today(), break_start_time))

                if annexe2_visit_to_hq_check:
                    messagebox.showwarning(title="Alerte",
                                           message=f"L'employé(e) ANNEXE-2 s'est rendu aU SIEGE entre: \n\n{annexe2_to_hq_entry} et {annexe2_to_hq_exit}")
                    # total_time = datetime.timedelta(hours=17, minutes=30)
                    if annexe2_to_hq_exit > end_time:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) - (
                                             datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                             datetime.datetime.combine(datetime.date.today(),
                                                                       break_start_time)) + datetime.datetime.combine(
                            datetime.date.today(),
                            annexe2_to_hq_exit) - datetime.datetime.combine(
                            datetime.date.today(), annexe2_to_hq_entry)

                    elif end_time > annexe2_to_hq_exit:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) - (
                                             datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                             datetime.datetime.combine(datetime.date.today(), break_start_time))

                if annexe2_visit_to_annexe1_check:
                    messagebox.showwarning(title="Alerte",
                                           message=f"L'employé(e) ANNEXE-2 s'est rendu a l'ANNEXE-1 entre: \n\n{annexe2_to_annexe1_entry} et {annexe2_to_annexe1_exit}")
                    # total_time = datetime.timedelta(hours=17, minutes=30)
                    if annexe2_to_annexe1_exit >= end_time:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) - (
                                             datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                             datetime.datetime.combine(datetime.date.today(),
                                                                       break_start_time)) + datetime.datetime.combine(
                            datetime.date.today(),
                            annexe2_to_annexe1_exit) - datetime.datetime.combine(
                            datetime.date.today(), annexe2_to_annexe1_entry)

                    elif end_time > annexe2_to_annexe1_exit:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) - (
                                             datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                             datetime.datetime.combine(datetime.date.today(), break_start_time))

                if work_case_exit:
                    messagebox.showwarning(title="Alerte",
                                           message=f"Sortie Signalée: Cadre du Travail / Autorisée entre: \n\n{new_entry} et {new_exit}")
                    # total_time = datetime.timedelta(hours=17, minutes=30)
                    if new_exit > end_time:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(),
                            start_time)) - (
                                             datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                             datetime.datetime.combine(datetime.date.today(), break_start_time)) + (
                                             datetime.datetime.combine(datetime.date.today(),
                                                                       new_exit) -
                                             datetime.datetime.combine(datetime.date.today(), new_entry))

                    elif end_time > new_exit:
                        total_time = (datetime.datetime.combine(datetime.date.today(),
                                                                end_time) - datetime.datetime.combine(
                            datetime.date.today(), start_time)) - (
                                             datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                             datetime.datetime.combine(datetime.date.today(), break_start_time))

                if personal_case_exit:
                    messagebox.showwarning(title="Alerte",
                                           message=f"Sortie Signalée: Hors Cadre du Travail entre: \n\n{personal_entry} et {personal_exit}")
                    total_time = (datetime.datetime.combine(datetime.date.today(),
                                                            end_time) - datetime.datetime.combine(
                        datetime.date.today(), start_time)) - (
                                         datetime.datetime.combine(datetime.date.today(), break_end_time) -
                                         datetime.datetime.combine(datetime.date.today(),
                                                                   break_start_time)) - (
                                         datetime.datetime.combine(datetime.date.today(),
                                                                   personal_exit) -
                                         datetime.datetime.combine(datetime.date.today(), personal_entry))

            total_time_str = str(total_time)
            self.result_label.config(text=total_time_str)
            return total_time  # Return the total_time value

        except ValueError:
            messagebox.showerror(title="Erreur", message="Verifier Heure")
        except TypeError:
            messagebox.showerror(title="Erreur", message="Entrée Invalide: Verifier Heure")

            ################################

    ### display the time spent by an employee at a different site
    def display_duration(self):

        """
        :return: The duration of the time spent by an employee
        """
        global duration_time

        # Break variables
        break_start_time_str = self.break_start_entry.get()  # fixed break start time
        break_end_time_str = self.break_end_entry.get()  # fixed break end time
        break_taken = self.break_check_button_var.get()

        break_start_time = datetime.datetime.strptime(break_start_time_str, "%H:%M").time()
        break_end_time = datetime.datetime.strptime(break_end_time_str, "%H:%M").time()

        # TEAM HQ TO (ANNEXE"1-2")
        hq_to_annexe1_entry_str = self.site_entry.get()
        hq_to_annexe1_exit_str = self.site_exit.get()
        hq_to_annexe2_entry_str = self.site_entry_01.get()
        hq_to_annexe2_exit_str = self.site_exit_01.get()

        # Conditional Verification Team HQ
        hq_to_annexe1_entry = datetime.datetime.strptime(hq_to_annexe1_entry_str, "%H:%M").time()
        hq_to_annexe1_exit = datetime.datetime.strptime(hq_to_annexe1_exit_str, "%H:%M").time()
        hq_to_annexe2_entry = datetime.datetime.strptime(hq_to_annexe2_entry_str, "%H:%M").time()
        hq_to_annexe2_exit = datetime.datetime.strptime(hq_to_annexe2_exit_str, "%H:%M").time()
        hq_visit_to_annexe1_check = self.exit_entry_status_var_1.get()
        hq_visit_to_annexe2_check = self.exit_entry_status_var_2.get()

        # Team ANNEXE 1 TO (HQ-ANNEXE 2)
        annexe1_to_hq_entry_str = self.annexe_entry.get()
        annexe1_to_hq_exit_str = self.annexe_exit.get()
        annexe1_to_annexe2_entry_str = self.annexe_entry_01.get()
        annexe1_to_annexe2_exit_str = self.annexe_exit_01.get()

        # Conditional Verification Team annexe1
        annexe1_to_hq_entry = datetime.datetime.strptime(annexe1_to_hq_entry_str, "%H:%M").time()
        annexe1_to_hq_exit = datetime.datetime.strptime(annexe1_to_hq_exit_str, "%H:%M").time()
        annexe1_to_annexe2_entry = datetime.datetime.strptime(annexe1_to_annexe2_entry_str, "%H:%M").time()
        annexe1_to_annexe2_exit = datetime.datetime.strptime(annexe1_to_annexe2_exit_str, "%H:%M").time()
        annexe1_visit_to_hq_check = self.presence_check_var.get()
        annexe1_visit_to_annexe2_check = self.annexe_to_annexe_var.get()

        # TEAM ANNEXE 2 TO (HQ-ANNEXE 1)
        annexe2_to_hq_entry_str = self.value_entry_widget.get()
        annexe2_to_hq_exit_str = self.value_exit_widget.get()
        annexe2_to_annexe1_entry_str = self.second_annexe_entry.get()
        annexe2_to_annexe1_exit_str = self.second_annexe_exit.get()

        # Conditional Verification Team Annexe2
        annexe2_to_hq_entry = datetime.datetime.strptime(annexe2_to_hq_entry_str, "%H:%M").time()
        annexe2_to_hq_exit = datetime.datetime.strptime(annexe2_to_hq_exit_str, "%H:%M").time()
        annexe2_to_annexe1_entry = datetime.datetime.strptime(annexe2_to_annexe1_entry_str, "%H:%M").time()
        annexe2_to_annexe1_exit = datetime.datetime.strptime(annexe2_to_annexe1_exit_str, "%H:%M").time()
        annexe2_visit_to_hq_check = self.verification_button_var.get()
        annexe2_visit_to_annexe1_check = self.second_verification_check_var.get()

        ### Complementary Entry / Exit
        new_entry_str = self.new_entry.get()
        second_exit_str = self.new_exit.get()
        personal_entry_str = self.personal_entry.get()
        personal_exit_str = self.personal_exit.get()

        #
        new_entry = datetime.datetime.strptime(new_entry_str, "%H:%M").time()
        new_exit = datetime.datetime.strptime(second_exit_str, "%H:%M").time()
        personal_entry = datetime.datetime.strptime(personal_entry_str, "%H:%M").time()
        personal_exit = datetime.datetime.strptime(personal_exit_str, "%H:%M").time()
        work_case_exit = self.first_btn_check_var.get()
        personal_case_exit = self.second_btn_check_var.get()

        ###

        if hq_visit_to_annexe1_check:
            duration_time = datetime.datetime.combine(datetime.date.today(),
                                                      hq_to_annexe1_exit) - datetime.datetime.combine(
                datetime.date.today(), hq_to_annexe1_entry)

        if hq_visit_to_annexe2_check:
            duration_time = datetime.datetime.combine(datetime.date.today(),
                                                      hq_to_annexe2_exit) - datetime.datetime.combine(
                datetime.date.today(), hq_to_annexe2_entry)

        if annexe1_visit_to_hq_check:
            duration_time = datetime.datetime.combine(datetime.date.today(),
                                                      annexe1_to_hq_exit) - datetime.datetime.combine(
                datetime.date.today(), annexe1_to_hq_entry)

        if annexe1_visit_to_annexe2_check:
            duration_time = datetime.datetime.combine(datetime.date.today(),
                                                      annexe1_to_annexe2_exit) - datetime.datetime.combine(
                datetime.date.today(), annexe1_to_annexe2_entry)

        if annexe2_visit_to_hq_check:
            duration_time = datetime.datetime.combine(datetime.date.today(),
                                                      annexe2_to_hq_exit) - datetime.datetime.combine(
                datetime.date.today(), annexe2_to_hq_entry)

        if annexe2_visit_to_annexe1_check:
            duration_time = datetime.datetime.combine(datetime.date.today(),
                                                      annexe2_to_annexe1_exit) - datetime.datetime.combine(
                datetime.date.today(), annexe2_to_annexe1_entry)

        if work_case_exit:
            duration_time = datetime.datetime.combine(datetime.date.today(),
                                                      new_exit) - datetime.datetime.combine(datetime.date.today(),
                                                                                            new_entry)

        if personal_case_exit:
            duration_time = datetime.datetime.combine(datetime.date.today(),
                                                      personal_exit) - datetime.datetime.combine(datetime.date.today(),
                                                                                                 personal_entry)

        if break_taken:
            duration_time = datetime.datetime.combine(datetime.date.today(),
                                                      break_end_time) - datetime.datetime.combine(
                datetime.date.today(), break_start_time)

        duration_time_str = str(duration_time)
        self.display_duration_label.config(text=duration_time_str)
        return duration_time

    ##
    ############## Excel File Generator function ###############

    # Excel file generator
    def save_to_excel(self):
        """
        Allows saving all required inputs through the GUI
        :return: Save information to an Excel file
        """
        nom_prenom = self.first_last_name_entry.get()
        fonction = self.title_combox.get()
        departement = self.department_combobox.get()
        arrivee = self.time_start_entry.get()
        # pause = self.break_check_button_var.get()
        debut_pause = self.break_start_entry.get()
        retour_pause = self.break_end_entry.get()
        descente = self.time_end_entry.get()
        lieu = self.place_combobox.get()
        total = self.calculate_total_time()
        jour_semaine = self.week_combobox.get()
        observation = self.observation_list_combobox.get()
        daily_date = self.date_entry.get()
        # duration = self.display_duration()
        standard_display = self.standard_duration_entry.get()

        # Validate input
        if not (nom_prenom and arrivee and descente and total):
            required_list = ["Nom & Prenom", "Heure Arrivee", "Heure Descente", "Total"]
            messagebox.showerror(f"Erreur: Sauvegarde-Archive Impossible",
                                 f"Veuillez remplir tout les champs requis:\n\n{list(required_list)}")
            return

        # Save data to Excel file
        try:
            file_path = f"Archive-Employee\\Archive_Employee_{nom_prenom}.xlsx"

            if not os.path.exists(file_path):
                # Create a workbook if it does not exist
                workbook = openpyxl.Workbook()
                workbook.iso_dates = True
                # sheet = workbook.active
                #
                #
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]

                    # Edit print options
                    sheet.print_options_horizontalCentered = True
                    sheet.print_options_verticalCentered = True

                    # Edit Page layout and size
                    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
                    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

                    # Set the print options to fit all columns on one-page
                    sheet.print_options.fitToWidth = True
                    sheet.print_options.fitToHeight = True
                    #
                    # # Set the scaling options
                    sheet.sheet_properties.pageSetUpPr.fitToPage = True
                    sheet.sheet_properties.pageSetUpPr.fitToWidth = 1
                    sheet.sheet_properties.pageSetUpPr.fitToHeight = 1

                    ### Default sheet column dimensions
                    column_widths = {
                        'A': 40, 'B': 33, 'C': 18, 'D': 11, 'E': 11, 'F': 9, 'G': 16, 'H': 16, 'I': 12, 'J': 40,
                        'K': 32, 'L': 10, 'M': 19}
                    for column, width in column_widths.items():
                        sheet.column_dimensions[column].width = width
                    #
                    # Row dimensions
                    row = sheet.row_dimensions[1]
                    row.height = 150

                    row = sheet.row_dimensions[2]
                    row.height = 25

                    row = sheet.row_dimensions[3]
                    row.height = 23

                    # Add a header Image to the Excel file
                    img_file = "images\\hci3n.png"
                    sheet.merge_cells('A1:E1')
                    img = Image(img_file)
                    img.width = 780
                    img.height = 185

                    first_cell = sheet['A1']
                    first_cell.alignment = Alignment(horizontal='center', vertical='center', mergeCell=True)
                    # Add the image to the worksheet
                    sheet.add_image(img)

                    # Sheet Dimensions
                    sheet.merge_cells('A2:M2')
                    header_value = 'Temps de Travail du Lundi au Jeudi: 8h-17h30--Total: 9h30/jour |<>|  Pause: 45min/jour |<>|  ' \
                                   'Vendredi: 8h-13h--Total: 5h  |<>| Total Semaine: 40h'
                    second_cell = sheet['A2']
                    second_cell.value = header_value
                    second_cell.fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0",
                                                   fill_type='lightTrellis')
                    second_cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Define the Excel sheet fill color
                    color1 = "00FF0000"  # red color
                    color2 = "0000CCFF"  # lightblue color
                    color3 = "00CCFFCC"  # lightgreen color
                    color4 = "00FF6600"  # orange color
                    color5 = "0000FF00"  # green color
                    color6 = "00C0C0C0"  # lightgrey

                    sheet.merge_cells('F1:M1')
                    sheet.title = f"Archive"
                    header_values = "SYNTHESES DES HORAIRES DE SERVICE DES EMPLOYÉ(ES) DU HC3N"
                    top_left_cell = sheet['F1']
                    top_left_cell.value = header_values
                    top_left_cell.fill = PatternFill(start_color=color6, end_color=color6, fill_type='lightTrellis')
                    top_left_cell.alignment = Alignment(horizontal='center', vertical='center')
                    list_to_append = ["NOM & PRENOM", "FONCTION", "DEPARTEMENT", "LIEU", "JOUR", "ENTREE",
                                      "DEBUT PAUSE",
                                      "RETOUR PAUSE", "DESCENTE", "TOTAL HEURE TRAVAIL EFFECTUER/Jr", "TOTAL HEURE A EFFECTUER/Jr", "DATE",
                                      "OBSERVATION"]
                    sheet.append(list_to_append)

                    # Save and close the workbook
                    ft = Font(bold=True, size=13)
                    fta = Font(bold=True, size=17)
                    border = Border(left=Side(border_style='thin', color='00000000'),
                                    right=Side(border_style='thin', color='00000000'),
                                    top=Side(border_style='thin', color='00000000'),
                                    bottom=(Side(border_style='thin', color='00000000')))

                    for row in sheet["A3:M3"]:
                        for cell in row:
                            cell.font = ft
                            cell.border = border

                    for row in sheet["F1:M1"]:
                        for cell in row:
                            cell.font = fta
                            cell.border = border

                    for row in sheet["A2:M2"]:
                        for cell in row:
                            cell.font = fta
                            cell.border = border

                    sheet["A3"].fill = PatternFill(start_color=color1, end_color=color1, fill_type='lightTrellis')
                    sheet["B3"].fill = PatternFill(start_color=color1, end_color=color1, fill_type='lightTrellis')
                    sheet["C3"].fill = PatternFill(start_color=color1, end_color=color1, fill_type='lightTrellis')
                    sheet["D3"].fill = PatternFill(start_color=color1, end_color=color1, fill_type='lightTrellis')
                    sheet["E3"].fill = PatternFill(start_color=color2, end_color=color2, fill_type='lightTrellis')
                    sheet["F3"].fill = PatternFill(start_color=color5, end_color=color5, fill_type='lightTrellis')
                    sheet["G3"].fill = PatternFill(start_color=color4, end_color=color4, fill_type='lightTrellis')
                    sheet["H3"].fill = PatternFill(start_color=color4, end_color=color4, fill_type='lightTrellis')
                    sheet["I3"].fill = PatternFill(start_color=color1, end_color=color1, fill_type='lightTrellis')
                    sheet["J3"].fill = PatternFill(start_color=color3, end_color=color3, fill_type='lightTrellis')
                    sheet["K3"].fill = PatternFill(start_color=color2, end_color=color2, fill_type='lightTrellis')
                    sheet["L3"].fill = PatternFill(start_color=color2, end_color=color2, fill_type='lightTrellis')
                    sheet["M3"].fill = PatternFill(start_color=color2, end_color=color2, fill_type='lightTrellis')

                # Save and close the workbook
                workbook.save(file_path)
                workbook.close()

            # Load the workbook
            workbook = load_workbook(file_path)
            sheet = workbook.active

            sheet.append([nom_prenom, fonction, departement, lieu, jour_semaine, arrivee, debut_pause, retour_pause,
                          descente, total, standard_display, daily_date, observation])
            # Save and close the workbook
            workbook.save(file_path)
            workbook.close()

            # Message to display after Success
            messagebox.showinfo(title="Succès", message="Donnée enregistrée avec succès !!!")

        # Exception message
        except Exception as e:
            messagebox.showerror("Error", str(e))

    ### Running function
    def run(self):
        self.window.mainloop()


#

if __name__ == "__main__":
    gui = JobTimeCalculator()
    gui.run()
